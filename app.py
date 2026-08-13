
from flask import Flask, render_template, jsonify
from pathlib import Path
import openpyxl
import os, re

app = Flask(__name__)
EXCEL_FILE = Path(os.environ.get("EXCEL_FILE", "VILLAREAL PAGINA.xlsm"))

def drive_image(url):
    """Convert common public Google Drive links to a direct thumbnail URL."""
    if not url:
        return ""
    url = str(url).strip()
    if not url:
        return ""
    patterns = [
        r"[?&]id=([A-Za-z0-9_-]+)",
        r"/file/d/([A-Za-z0-9_-]+)",
        r"/d/([A-Za-z0-9_-]+)"
    ]
    if "drive.google.com" in url:
        for pat in patterns:
            m = re.search(pat, url)
            if m:
                return f"https://drive.google.com/thumbnail?id={m.group(1)}&sz=w1800"
    if url.startswith(("http://", "https://", "/")):
        return url
    filename = Path(url.replace("\\", "/")).name
    return f"/static/images/{filename}" if filename else ""

def read_catalog():
    if not EXCEL_FILE.exists():
        return []
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True, keep_vba=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(x or "").strip().upper() for x in rows[0]]

    items = []
    for row in rows[1:]:
        data = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
        if not data.get("PRODUCTO"):
            continue
        visible = str(data.get("VISIBLE") or "").strip().upper()
        if visible in {"NO", "0", "FALSE"}:
            continue

        image_keys = ["IMAGEN_URL", "MAGEN_URL", "IMAGEN_URL_2", "IMAGEN_URL_3", "IMAGEN_URL_4"]
        images = []
        for k in image_keys:
            img = drive_image(data.get(k))
            if img and img not in images:
                images.append(img)

        items.append({
            "codigo": data.get("CODIGO", ""),
            "producto": data.get("PRODUCTO", ""),
            "categoria": data.get("CATEGORIA", ""),
            "medida": data.get("MEDIDA", ""),
            "color": data.get("COLOR", ""),
            "acabado": data.get("ACABADO", ""),
            "precio": data.get("PRECIO", ""),
            "descripcion": data.get("DESCRIPCION", ""),
            "images": images
        })
    return items

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/productos")
def productos():
    return jsonify(read_catalog())

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
